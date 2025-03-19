import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import openpyxl.utils
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# マークダウンファイルからExcelファイルを作成する
def markdown_to_excel(md_file_path, excel_file_path):
    ### Excelファイル作成
    wb = openpyxl.Workbook()
    ### シートの作成し、アクティブシートにする。
    ws = wb.active
    ### シートのタイトルをマークダウンファイル名に変更する
    ws.title = "構築手順書"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

	# 列の幅の調整
    column_width = [1, 4, 23, 4, 21, 60, 5, 12, 12, 12, 8, 12, 12, 40]
    for i, width in enumerate(column_width, start=1):
        # よくわからない…
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 1行目に作業証跡凡例を記載
    exp = ws["B1"]
    exp.value = "作業証跡取得欄 ○:画面のハードコピー ●:手順書指定の証跡取得 ー:不要"
    exp.alignment = Alignment(horizontal="left", vertical="center")

    # 2行目に「作業手順詳細」行を追加する
    detail = ws["B2"]
    detail.value = "作業手順詳細"

    # 「作業手順詳細」を外枠で囲む
    for col in range(2, 15):
         cell = ws.cell(row=2, column=col)
         cell.border = thin_border

    
    # 「作業手順詳細」の塗りつぶし
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=14):
        for cell in row:
            cell.fill = PatternFill(start_color='0099CCFF', fill_type='solid')
    
    # 3行目にヘッダーを作成する
    headers = ["No", "作業項目", "No", "作業対象", "作業手順", "証跡\n取得", "チェック", "作業者", "作業日", "想定\n時間", "開始\n時刻", "完了\n時刻", "備考"]
    ws.append([None] + headers)
    for col_num, header in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.fill = PatternFill(start_color='0099CCFF', fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
 
    # マークダウンファイルの読み取り
    with open (markdown_file, "r",encoding="utf-8") as f:
        lines = f.readlines()

    mn = 0
    bn = 0
    mi = di = bi = ""

    for line in lines:
        line = line.strip()

        # 作業項目の処理
        if line.startswith("### "):
            mn += 1
            bn = 0
            mi = line[4:].strip()
            ws.append([None, mn, mi])

        # 作業手順の処理
        elif line.startswith("- "):
            bn += 1
            bi = line[2:].strip()
            ws.append([None, None, None, bn, None, bi])

    # B3からN列まで枠線を適用し、格子状に囲む
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=3, max_row = max_row, min_col=2,max_col=14):
            for cell in row:
                    cell.border = thin_border

    #J列を日付表記にする
    for row in range(4, max_row + 1):
        cell = ws.cell(row = row, column = 10) # J列
        cell.number_format = "mm/dd"
							
    # L列とM列を時刻表記にする
    for row in range(4, max_row + 1):
        cell = ws.cell(row = row, column = 12) # L列
        cell.number_format = "hh:mm"
				
    for row in range(4, max_row + 1):
        cell = ws.cell(row = row, column = 13) # M列
        cell.number_format = "hh:mm"
				
	# G列にラジオボタンの選択肢を設定
    dv = DataValidation(type="list",formula1 = '"ー,○,●"',showDropDown = True)
    ws.add_data_validation(dv)
    for row in range(4, max_row + 1):
        cell = ws.cell(row=row, column=7) # G列
        dv.add(cell)
        cell.alignment = Alignment(horizontal = "center", vertical = "center")
			
    # H列にチェックボックスを設定
    for row in range(4, max_row +1):
        cell = ws.cell(row=row, column=8) # H列		
        cell.value = "□"
        cell.alignment = Alignment(horizontal = "center", vertical = "center")
 
    # 折り返して全体を表示する
    wrap_text= Alignment(wrapText=True)
    for row in ws.iter_rows(min_row=4, min_col=3,max_col=3):
         for cell in row:
              cell.alignment = wrap_text

    for row in ws.iter_rows(min_row=4, min_col=6,max_col=6):
        for cell in row:
            cell.alignment = wrap_text  
 
     # フォントの設定
    font = Font(name="Meiryo UI")
    for row in ws:
         for cell in row:
              cell.font = font
    
    for cell in ws.iter_cols(min_col=2, max_col=14, min_row=2, max_row=3):
             for c in cell:
                  c.font = Font(name="Meiryo UI", bold=True)

    # セルの結合
    def merge_same_cells(ws, col, start_row):
        max_row = ws.max_row
        merge_start = start_row # 結合開始位置
        prev_value = ws[f"{col}{start_row}"].value #最初の値

        for row in range(start_row + 1, max_row + 1):
            current_value = ws[f"{col}{row}"].value

            if current_value is not None: # 値がある場合のみ、新しい結合を開始
                if merge_start < row - 1: # 2行以上の場合のみ結合
                    ws.merge_cells(f"{col}{merge_start}:{col}{row - 1}")
                    ws[f"{col}{merge_start}"].alignment = Alignment(vertical="center")

            # 次の結合開始位置を更新
            merge_start = row
            prev_value = current_value

        if merge_start < max_row:
            ws.merge_cells(f"{col}{merge_start}:{col}{max_row}")
            ws[f"{col}{merge_start}"].alignment = Alignment(vertical="center")
    
    merge_same_cells(ws, col="B", start_row =4)

    # Excelファイルの保存
    wb.save(excel_file_path)

# 出力
markdown_file = "見本.md"  # 入力マークダウンファイルのパス
excel_file = "見本.xlsx"    # 出力Excelファイルのパス
markdown_to_excel(markdown_file, excel_file)